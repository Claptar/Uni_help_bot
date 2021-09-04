import os.path
import pickle
import time

import pandas as pd
from openpyxl.utils import datetime
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles.colors import COLOR_INDEX

from data_constructor import psg


def within_range(bounds: tuple, cell: Cell) -> bool:
    """
    Функция, определяющая, входит ли клетка в состав большой слитой или нет.
    :param bounds: границы merged клеток
    :param cell: сама клетка
    :return: True, если merged клетка, иначе False
    """
    column_start, row_start, column_end, row_end = bounds  # границы merged клетки
    row = cell.row  # проверка, находится ли клетка в этом слиянии
    if row_start <= row <= row_end:  # ___________________
        column = cell.column  # |value|empty|empty|
        if (
            column_start <= column <= column_end
        ):  # |empty|empty|empty|  Пример merged клетки
            return True  # |empty|empty|empty|
    return False  #


def get_value_merged(sheet: Worksheet, cell: Cell) -> any:
    """
    Функция, возвращающая значение, лежащее в клетке, вне зависимости от того, является ли клетка merged, или нет.
    :param sheet: таблица с расписанием
    :param cell: клетка таблицы
    :return: значение, лежащее в клетке
    """
    for (
        merged
    ) in (
        sheet.merged_cells
    ):  # смотрим в списке слитых клеток (структура данных openpyxl.worksheet)
        if within_range(merged.bounds, cell):
            return sheet.cell(
                merged.min_row, merged.min_col
            ).value  # смотрим значение в левой верхней клетке
    return cell.value


def get_color_merged(sheet: Worksheet, cell: Cell) -> any:
    """
    Функция, возвращающая цвет клетки, вне зависимости от того, является ли клетка merged, или нет.
    :param sheet: таблица с расписанием
    :param cell: клетка таблицы
    :return: значение, лежащее в клетке
    """
    for (
        merged
    ) in (
        sheet.merged_cells
    ):  # смотрим в списке слитых клеток (структура данных openpyxl.worksheet)
        if within_range(merged.bounds, cell):
            # смотрим цвет левой верхней клетки
            color = sheet.cell(merged.min_row, merged.min_col).fill.start_color.index
            color = (
                "#" + COLOR_INDEX[color][2:] if type(color) == int else "#" + color[2:]
            )
            return color
    color = cell.fill.start_color.index
    color = "#" + COLOR_INDEX[color][2:] if type(color) == int else "#" + color[2:]
    return color


def insert_update_group_timetable(group_name, timetable, exam=False):
    """
    Функция, вставляющая или обновляющая значение расписания в таблице.
    :param group_name: номер группы
    :param timetable: расписание группы
    :param exam: если True, то обновляется расписание экзаменов группы
    :return:
    """
    insert = psg.sync_insert_group(
        group_name, pickle.dumps(timetable, protocol=pickle.HIGHEST_PROTOCOL), exam=exam
    )
    timeout = (
        time.time() + 30
    )  # если подключение к серверу длится дольше 30 секунд, то вызываем ошибку
    while not insert[0] and insert[1] == "connection_error":
        insert = psg.sync_insert_group(
            group_name,
            pickle.dumps(timetable, protocol=pickle.HIGHEST_PROTOCOL),
            exam=exam,
        )
        if time.time() > timeout:
            raise RuntimeError
    # если возникает какая-то ошибка при вставке, то пробуем обновить значение, а не вставить
    if not insert[0] and insert[1] == "other_error":
        update = psg.sync_update_group(
            group_name,
            pickle.dumps(timetable, protocol=pickle.HIGHEST_PROTOCOL),
            exam=exam,
        )
        timeout = (
            time.time() + 30
        )  # если подключение к серверу длится дольше 30 секунд, то вызываем ошибку
        while not update[0] and update[1] == "connection_error":
            update = psg.sync_update_group(
                group_name,
                pickle.dumps(timetable, protocol=pickle.HIGHEST_PROTOCOL),
                exam=exam,
            )
            if time.time() > timeout:
                raise RuntimeError


def get_timetable(table: Worksheet):
    """
    Функция, которая из таблицы Excel с расписанием выделяет расписание для каждой группы
    и записывает его в базу данных.
    :param table: таблица с расписанием
    :return:
    """
    hours_list = [
        "09:00 – 10:25",
        "10:45 – 12:10",
        "12:20 – 13:45",
        "13:55 – 15:20",
        "15:30 – 16:55",
        "17:05 – 18:30",
        "18:35 - 20:00",
    ]
    alumni_timetable = None
    for j in range(3, table.max_column + 1):  # смотрим на значения по столбцам
        group_name = table.cell(9, j).value  # номер группы
        if group_name in [
            "Дни",
            "Часы",
        ]:  # если это не номер группы, то пропускаем столбец
            continue
        # иначе если столбец - это номер группы, то составляем для него расписание
        elif group_name is not None:
            if isinstance(
                group_name, int
            ):  # если номер группы - просто число, преобразуем его в строку
                group_name = str(group_name)
            # group - словарь с расписанием для группы
            timetable = dict(
                Понедельник={},
                Вторник={},
                Среда={},
                Четверг={},
                Пятница={},
                Суббота={},
                Воскресенье={},
            )
            for k in range(10, table.max_row + 1):  # проходимся по столбцу
                # если клетки относятся ко дню недели (не разделители)
                if get_value_merged(table, table.cell(k, 1)) in timetable:
                    day = get_value_merged(
                        table, table.cell(k, 1)
                    )  # значение дня недели
                    hours = get_value_merged(
                        table, table.cell(k, 2)
                    )  # клетка, в которой лежит значение времени
                    pair = get_value_merged(
                        table, table.cell(k, j)
                    )  # клетка, в которой лежит значение пары
                    color = get_color_merged(table, table.cell(k, j))  # цвет клетки
                    # цветные круги, соответствующие семинарам, лабам / англу, лекциям, базовому дню и военке
                    colors_to_circles = {
                        "#CCFFFF": "🔵",  # семинары
                        "#92D050": "🔵",  # семинары
                        "#00FFFF": "🔵",  # семинары
                        "#66FFFF": "🔵",  # семинары
                        "#FFFFFF": "🔵",  # семинары
                        "#00B050": "🔵",  # семинары
                        "#FFFF99": "🟡",  # лабы / англ
                        "#FF99CC": "🔴",  # лекции
                        "#CCFFCC": "🟢",  # базовый день
                        "#FFC000": "🟠",  # военка
                    }

                    # рассматриваем только те клетки, для которых определено значение как пары, так и времени
                    if hours is not None and pair is not None:
                        hours = (
                            hours.split()
                        )  # преобразуем время пары к формату hh:mm – hh:mm
                        if len(hours[0][:-2]) == 1:
                            hours[0] = "0" + hours[0]
                        hours = (
                            hours[0][:-2]
                            + ":"
                            + hours[0][-2:]
                            + " – "
                            + hours[2][:-2]
                            + ":"
                            + hours[2][-2:]
                        )
                        # записываем значение в расписание
                        try:
                            timetable[day][hours] = (
                                colors_to_circles[color] + " " + pair
                                if pair is not None
                                else pair
                            )
                        except KeyError:  # если появится новый цвет, то он будет выведен на экран
                            print(color, pair)
            timetable = pd.DataFrame(
                timetable, columns=timetable.keys(), index=hours_list, dtype=object
            )
            timetable.replace(
                to_replace=[None], value="😴", inplace=True
            )  # заменяем None на спящие смайлики
            # на первой итерации записываем пустую табличку для выпускников (если нужно)
            if (
                not os.path.exists("semester/blank_timetable.pickle")
                and alumni_timetable is None
            ):
                alumni_timetable = timetable

            # записываем или обновляем номер группы и расписание в базу данных
            insert_update_group_timetable(group_name, timetable)
    # записываем или обновляем расписание для выпускников
    if alumni_timetable is not None:
        alumni_timetable.iloc[:] = "😴"
        with open("semester/blank_timetable.pickle", "wb") as handle:
            pickle.dump(alumni_timetable, handle, protocol=pickle.HIGHEST_PROTOCOL)


def get_exam_timetable(table: Worksheet):
    """
    Функция, которая из таблицы Excel с расписанием экзаменов выделяет расписание для каждой группы
    и записывает его в базу данных.
    :param table: таблица с расписанием
    :return:
    """
    for j in range(3, table.max_column + 1):  # смотрим на значения по столбцам
        group_name = table.cell(7, j).value  # номер группы
        if group_name is not None:
            if isinstance(
                group_name, int
            ):  # если номер группы - просто число, преобразуем его в строку
                group_name = str(group_name)
            # group - словарь с расписанием для группы
            group_name = "".join(group_name.split())
            timetable = dict(Экзамены={})
            for k in range(8, table.max_row + 1):  # проходимся по столбцу
                # если клетки относятся ко дню недели (не разделители)
                date = get_value_merged(table, table.cell(k, 2))  # значение дня
                week_day = get_value_merged(table, table.cell(k, 1))  # день недели
                if date is not None:
                    if date.month == 12:
                        month = "декабря"
                    elif date.month == 1:
                        month = "января"
                    elif date.month == 5:
                        month = "мая"
                    elif date.month == 6:
                        month = "июня"
                    else:
                        month = ""
                    day = (
                        str(date.day) + " " + month + " " + "(" + week_day.lower() + ")"
                    )
                    exam = get_value_merged(
                        table, table.cell(k, j)
                    )  # клетка, в которой лежит значение пары
                    if exam is not None:
                        timetable["Экзамены"][day] = exam
                else:
                    continue
            try:
                timetable = pd.DataFrame(
                    timetable, dtype=object, index=timetable["Экзамены"].keys()
                )
            except TypeError:
                print(timetable)
            insert_update_group_timetable(group_name, timetable, exam=True)
